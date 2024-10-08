import openpyxl
import requests
import os
from dotenv import load_dotenv

load_dotenv()

ERROR_MESSAGE_NOT_SET_ADDRESS = "職員住所または勤務先が未設定"

ERROR_FILE_TITLE = ["職員番号", "氏名", "エラーメッセージ"]


def search_address(adress):
    url = os.environ["SEARCH_URL"]
    key = os.environ["API"]
    host = os.environ["SEARCH_HOST"]

    querystring = {
        "addr": adress,
        "gov": "0",
        "fmt": "json",
    }

    headers = {
        "x-rapidapi-key": f"{key}",
        "x-rapidapi-host": f"{host}",
    }

    response = requests.get(url, headers=headers, params=querystring)

    dic = response.json()

    search_remaining = response.headers["X-RateLimit-Requests-Remaining"]

    lon = dic["results"][0]["lon"]
    lat = dic["results"][0]["lat"]

    lon_lat = f"{lon},{lat}"

    return lon_lat, search_remaining


def search_route(start, destination):
    url = os.environ["ROUTE_URL"]
    key = os.environ["API"]
    host = os.environ["ROUTE_HOST"]

    querystring = {
        "start": start,
        "destination": destination,
        "priority": "0",
        "tollway": "0",
        "ferry": "0",
        "smartic": "0",
        "etc": "0",
        "tolltarget": "0",
        "cartype": "0",
        "vehicletype": "0",
        "danger": "0",
        "daytime": "0",
        "generalroad": "0",
        "tollroad": "0",
        "regulations": "0",
        "travel": "0",
        "uturnavoid": "0",
        "uturn": "0",
        "resulttype": "0",
        "fmt": "json",
    }

    headers = {
        "x-rapidapi-key": f"{key}",
        "x-rapidapi-host": f"{host}",
    }

    response = requests.get(url, headers=headers, params=querystring)

    dic = response.json()

    route_remaining = response.headers["X-RateLimit-Requests-Remaining"]

    # print(dic)

    distance = dic["summary"]["totalDistance"]

    return distance, route_remaining


def is_empty(value):
    return value is None


wb = openpyxl.load_workbook("C:/Users/pytho/OneDrive/Desktop/20241004/route_search_error1.xlsx")
ws = wb["Sheet1"]

count = 0
row_list = []
office_list = []
error_list = []
for row in ws.iter_rows(min_row=2, max_col=5):
    # 範囲の最小列が２、範囲の最大列が５
    if is_empty(row[0].value):
        continue

    row_list.append([cell.value for cell in row])

tmp_list = list(set([row[3] for row in row_list if not is_empty(row[3])]))

for office in tmp_list:
    lon_lat, search_remaining = search_address(office)
    office_list.append([office, lon_lat])

for row in row_list:
    staff_address = row[2]
    office_address = row[3]

    if is_empty(staff_address) or is_empty(office_address):
        error_list.append([row[0], row[1], ERROR_MESSAGE_NOT_SET_ADDRESS])
        continue

    staff_lon_lat, search_remaining = search_address(staff_address)
    # office_lon_lat, search_remaining = search_address(office_address)

    for office in office_list:
        if office_address == office[0]:
            office_lon_lat = office[1]
            break

    distance, route_remaining = search_route(staff_lon_lat, office_lon_lat)

    d = round(distance / 1000, 1)

    row[4] = d

for idx, row in enumerate(row_list):
    ws.cell(idx + 2, 5).value = row[4]

wb.save("route_search_error1.xlsx")


# 誰から見てもわかりにくい変数設定になってしまった

print(f"SearchAPIの残り回数:{search_remaining}")
print(f"RouteAPIの残り回数:{route_remaining}")

if error_list:
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.active

    write_ws.append(ERROR_FILE_TITLE)

    for idx_row, row in enumerate(error_list):
        for idx_col, col in enumerate(row):
            write_ws.cell(idx_row + 2, idx_col + 1).value = col

    write_wb.save("error_list.xlsx")
